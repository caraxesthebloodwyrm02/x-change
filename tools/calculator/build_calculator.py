#!/usr/bin/env python3
"""Build x-change Reward & Token Calculator spreadsheet."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── ARGB colour constants ─────────────────────────────────────────────────────
BLUE_TXT   = "FF0000FF"   # hardcoded inputs
BLACK_TXT  = "FF000000"   # formulas / labels
GREEN_TXT  = "FF008000"   # cross-sheet links
YELLOW_BG  = "FFFFFF00"   # key assumption cells
HDR_BG     = "FF1F4E79"   # dark-blue header background
HDR_FG     = "FFFFFFFF"   # white header text
SECT_BG    = "FFBDD7EE"   # section sub-header
ALT_BG     = "FFF2F2F2"   # alternating rows
PASS_BG    = "FFC6EFCE"   # green — layer pass
ERR_BG     = "FFFFC7CE"   # red — layer fail
WARN_BG    = "FFFFEB9C"   # orange — stripped / narrowed
TITLE_BG   = "FF244061"

FONT  = "Arial"
SZ    = 10
SZ_T  = 16
SZ_S  = 11


# ── Helper factories ──────────────────────────────────────────────────────────
def _f(color=BLACK_TXT, bold=False, sz=SZ, italic=False):
    return Font(name=FONT, size=sz, bold=bold, italic=italic, color=color)

def _fill(c): return PatternFill("solid", start_color=c)
def _left():  return Alignment(horizontal="left",   vertical="center", wrap_text=False)
def _ctr():   return Alignment(horizontal="center", vertical="center")
def _wrap():  return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _border():
    t = Side(style="thin", color="FFAAAAAA")
    return Border(left=t, right=t, top=t, bottom=t)

def _thick_bottom():
    thin  = Side(style="thin",  color="FFAAAAAA")
    thick = Side(style="medium", color="FF1F4E79")
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def w(ws, row, col, value=None, font=None, fill=None, align=None,
      border=None, fmt=None, height=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font      = font
    if fill:   c.fill      = fill
    if align:  c.alignment = align
    if border: c.border    = border
    if fmt:    c.number_format = fmt
    if height: ws.row_dimensions[row].height = height
    return c


def hdr_row(ws, row, labels, start_col=1, height=18):
    ws.row_dimensions[row].height = height
    for i, lbl in enumerate(labels):
        w(ws, row, start_col + i, lbl,
          font=_f(HDR_FG, bold=True), fill=_fill(HDR_BG),
          align=_ctr(), border=_border())


def sect(ws, row, col, text, span=1, height=16):
    ws.row_dimensions[row].height = height
    c = w(ws, row, col, text,
          font=_f(HDR_FG, bold=True, sz=SZ_S),
          fill=_fill(SECT_BG if "──" not in text else HDR_BG),
          align=_left(), border=_border())
    c.fill = _fill(HDR_BG)
    c.font = _f(HDR_FG, bold=True, sz=SZ_S)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return c


def inp(ws, row, col, value, fmt=None):
    return w(ws, row, col, value,
             font=_f(BLUE_TXT), fill=_fill(YELLOW_BG),
             align=_left(), border=_border(), fmt=fmt)


def form(ws, row, col, formula, fmt=None, green=False):
    return w(ws, row, col, formula,
             font=_f(GREEN_TXT if green else BLACK_TXT),
             align=_left(), border=_border(), fmt=fmt)


def lbl(ws, row, col, text, bold=True):
    return w(ws, row, col, text,
             font=_f(BLACK_TXT, bold=bold), align=_left())


def col_w(ws, mapping):
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


# ── TIER LOOKUP (shared data) ─────────────────────────────────────────────────
TIERS = [
    ("surface",     0.00, 1, "Fact recall, basic recognition"),
    ("pattern",     0.25, 2, "Cross-domain pattern recognition"),
    ("structural",  0.50, 3, "Structural understanding, system maps"),
    ("causal",      0.75, 4, "Causal chain modelling, root cause"),
    ("theoretical", 1.00, 5, "Novel synthesis, theory building"),
]
TIER_NAMES = [t[0] for t in TIERS]

STATES = [
    "drafted", "earned", "payment_pending",
    "payment_confirmed", "student_acknowledged", "review_requested",
]

EVIDENCE_TYPES = [
    "glass_session_event", "failure_snapshot", "return_attempt",
    "success_after_support", "agent_interpretation", "student_confirmation",
]


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Cover
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    # Title band
    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:F1")
    w(ws, 1, 1, "x-change  —  Reward & Token Calculator",
      font=_f(HDR_FG, bold=True, sz=SZ_T),
      fill=_fill(TITLE_BG), align=_ctr(), border=_border())

    ws.row_dimensions[2].height = 8

    # Subtitle / description
    ws.row_dimensions[3].height = 14
    ws.merge_cells("A3:F3")
    w(ws, 3, 1,
      "Policy source: docs/policy-core-v0.md  |  Code: src/xchange/domain.py  |  Version: 1.0",
      font=_f(BLACK_TXT, italic=True, sz=9), align=_left())

    ws.row_dimensions[4].height = 60
    ws.merge_cells("A4:F4")
    w(ws, 4, 1,
      ("This model implements the x-change v0 epistemic token lifecycle in full. "
       "It covers: (1) rarity scoring via the compute_rarity_score() formula with "
       "configurable weights, (2) all five InsightTier classifications and their "
       "backward-compatible integer amounts, (3) the six-state reward lifecycle and "
       "valid transition paths, (4) a live 15-token portfolio log with formula-driven "
       "rarity scores, (5) layered exchange constraint evaluation (safety → ethics → "
       "economic), and (6) a portfolio summary dashboard. "
       "All yellow cells are editable inputs (blue text). All black text cells are "
       "formulas — do not hardcode them."),
      font=_f(BLACK_TXT, sz=SZ), align=_wrap(), border=_border())

    ws.row_dimensions[5].height = 8

    # Navigation table
    sect(ws, 6, 1, "Navigation", span=4)
    hdr_row(ws, 7, ["Sheet", "Purpose", "Key Inputs", "Key Outputs"])
    nav = [
        ("Assumptions",   "All configurable parameters — edit here first",
         "Rarity weights, tier table, constraint config, blocked / irreversible key lists",
         "Named inputs referenced by all other sheets"),
        ("Rarity_Calc",   "Single-token rarity score calculator",
         "Insight tier, inferential richness, trend position",
         "Rarity score (0-1), component breakdown, tier amount"),
        ("Tier_Ref",      "InsightTier reference table",
         "None (reference only)",
         "Tier order, rarity weight, amount, description"),
        ("Lifecycle",     "Reward state machine and valid transitions",
         "Current state, evidence type, policy flags",
         "Valid next states, outcome state, notes"),
        ("Token_Log",     "Portfolio of 15 sample issued tokens",
         "Per-token: tier, base_bank_depth, inferential richness, trend position",
         "Rarity score, amount, portfolio-level stats"),
        ("Exchange_Eval", "Layered exchange constraint evaluator",
         "Requested scope (up to 10 key-value pairs), explicit approval flag",
         "Per-layer pass/block/strip result, final approved scope, overall decision"),
        ("Summary",       "Portfolio dashboard — aggregate view",
         "Driven by Token_Log (no manual input)",
         "Tokens by tier, avg rarity by tier, total amount, rarity bands"),
    ]
    for i, (sh, purp, inp_desc, out) in enumerate(nav):
        row = 8 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        for j, val in enumerate([sh, purp, inp_desc, out]):
            font = _f(GREEN_TXT if j == 0 else BLACK_TXT, bold=(j == 0))
            c = w(ws, row, 1 + j, val, font=font, align=_wrap(), border=_border())
            ws.row_dimensions[row].height = 30
            if row_fill: c.fill = row_fill

    ws.row_dimensions[16].height = 8
    sect(ws, 17, 1, "Model Metadata", span=4)
    meta = [
        ("Policy Version",    "policy-core-v0"),
        ("Model Version",     "1.0"),
        ("Domain Code",       "src/xchange/domain.py"),
        ("Rarity Formula",    "0.4×(1−trend_position) + 0.4×inferential_richness + 0.2×tier_weight"),
        ("Token Amounts",     "surface=1, pattern=2, structural=3, causal=4, theoretical=5"),
        ("Constraint Layers", "1. Safety/Security (hard gate)  2. Ethics/Irreversibility  3. Economic Stability"),
        ("Build Date",        "2026-05-07"),
    ]
    for i, (key, val) in enumerate(meta):
        row = 18 + i
        lbl(ws, row, 1, key)
        w(ws, row, 2, val, font=_f(BLACK_TXT), align=_left(), border=_border())

    col_w(ws, {"A": 18, "B": 35, "C": 42, "D": 42, "E": 12, "F": 12})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Assumptions
# ═══════════════════════════════════════════════════════════════════════════════
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    # ── A. Rarity score weights ───────────────────────────────────────────────
    sect(ws, 1, 1, "A. Rarity Score Weights", span=4)
    hdr_row(ws, 2, ["Parameter", "Value", "Validation", "Source"])
    rarity_params = [
        ("Emergence Weight (w_emergence)",            0.4),
        ("Inferential Richness Weight (w_inferential)",0.4),
        ("Epistemic Tier Weight (w_tier)",            0.2),
    ]
    for i, (label, val) in enumerate(rarity_params):
        row = 3 + i
        lbl(ws, row, 1, label)
        inp(ws, row, 2, val, fmt="0.0%")
    # sum check
    lbl(ws, 6, 1, "Weight Sum  (must equal 100.0%)")
    form(ws, 6, 2, "=SUM(B3:B5)", fmt="0.0%")
    form(ws, 6, 3, '=IF(ABS(B6-1)<0.0001,"✓ OK","✗ ERROR — weights must sum to 1.0")')
    ws.cell(row=3, column=4).value = "Source: domain.py:compute_rarity_score, line 111"
    ws.cell(row=3, column=4).font  = _f(BLACK_TXT, sz=9, italic=True)

    # ── B. Insight Tier Reference Table ──────────────────────────────────────
    sect(ws, 8,  1, "B. Insight Tier Reference Table", span=5)
    hdr_row(ws, 9, ["Tier (text key)", "Rarity Weight", "Amount (int)", "Description", "Tier Order"])
    for i, (tier, rw, amt, desc) in enumerate(TIERS):
        row = 10 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        inp(ws, row, 1, tier)
        c2 = inp(ws, row, 2, rw,  fmt="0.00")
        c3 = inp(ws, row, 3, amt)
        c4 = w(ws, row, 4, desc, font=_f(BLACK_TXT), align=_left(), border=_border())
        c5 = w(ws, row, 5, i+1,  font=_f(BLACK_TXT), align=_ctr(), border=_border())
        if row_fill:
            for c in (c4, c5): c.fill = row_fill
    ws.cell(row=15, column=4).value = "Source: domain.py:InsightTier + _TIER_AMOUNTS, lines 48-79"
    ws.cell(row=15, column=4).font  = _f(BLACK_TXT, sz=9, italic=True)

    # ── C. Economic Constraint Defaults ──────────────────────────────────────
    sect(ws, 17, 1, "C. Economic Constraint Defaults", span=4)
    hdr_row(ws, 18, ["Parameter", "Value", "", "Source"])
    econ_params = [
        ("Max Token Amount (cap on token_amount scope key)", 10,    "Source: domain.py:ConstraintConfig.max_token_amount"),
        ("Max Scope Items (max keys approved)",               5,     "Source: domain.py:ConstraintConfig.max_scope_items"),
        ("Require Explicit Irreversible Approval (TRUE/FALSE)", "TRUE", "Source: domain.py:ConstraintConfig.require_explicit_irreversible_approval"),
    ]
    for i, (label, val, src) in enumerate(econ_params):
        row = 19 + i
        lbl(ws, row, 1, label, bold=False)
        inp(ws, row, 2, val)
        ws.cell(row=row, column=4).value = src
        ws.cell(row=row, column=4).font  = _f(BLACK_TXT, sz=9, italic=True)

    # ── D. Blocked Scope Keys ─────────────────────────────────────────────────
    sect(ws, 23, 1, "D. Blocked Scope Keys  (Safety/Security — Hard Gate)", span=4)
    hdr_row(ws, 24, ["Scope Key", "Reason for Block", "", ""])
    blocked = [
        ("admin_override",      "Bypasses policy — never approvable"),
        ("raw_db_write",        "Direct database mutation; unsafe"),
        ("delete_reward",       "Irreversible data destruction"),
        ("bypass_constraints",  "Self-referential safety bypass"),
        ("payment_redirect",    "Financial fraud vector"),
    ]
    for i, (key, reason) in enumerate(blocked):
        row = 25 + i
        inp(ws, row, 1, key)
        w(ws, row, 2, reason, font=_f(BLACK_TXT), align=_left(), border=_border())
    # 3 blank extension rows
    for i in range(3):
        row = 30 + i
        inp(ws, row, 1, "")
        w(ws, row, 2, "(add custom blocked key here)", font=_f(BLACK_TXT, italic=True, sz=9),
          align=_left(), border=_border())

    # ── E. Irreversible Scope Keys ────────────────────────────────────────────
    sect(ws, 34, 1, "E. Irreversible Scope Keys  (Ethics/Irreversibility Layer)", span=4)
    hdr_row(ws, 35, ["Scope Key", "Reason", "", ""])
    irrev = [
        ("finalize_contract",  "Contract closure cannot be undone"),
        ("archive_student",    "Student record archival is permanent"),
        ("issue_final_token",  "Token issuance is immutable"),
        ("submit_stripe_charge","Payment charges non-reversible automatically"),
    ]
    for i, (key, reason) in enumerate(irrev):
        row = 36 + i
        inp(ws, row, 1, key)
        w(ws, row, 2, reason, font=_f(BLACK_TXT), align=_left(), border=_border())
    for i in range(3):
        row = 40 + i
        inp(ws, row, 1, "")
        w(ws, row, 2, "(add custom irreversible key here)", font=_f(BLACK_TXT, italic=True, sz=9),
          align=_left(), border=_border())

    col_w(ws, {"A": 48, "B": 40, "C": 30, "D": 60, "E": 14})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Rarity_Calc
# ═══════════════════════════════════════════════════════════════════════════════
def build_rarity_calc(wb):
    ws = wb.create_sheet("Rarity_Calc")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Single-Token Rarity Score Calculator", span=4)
    w(ws, 2, 1,
      "Enter inputs below (yellow cells). Scores are computed by formulas referencing Assumptions!",
      font=_f(BLACK_TXT, italic=True, sz=9), align=_wrap())
    ws.merge_cells("A2:D2")

    # ── Inputs ────────────────────────────────────────────────────────────────
    sect(ws, 4, 1, "Inputs", span=4)
    hdr_row(ws, 5, ["Field", "Value", "Allowed Range", "Notes"])
    inputs_def = [
        ("insight_tier",         "causal",  TIER_NAMES,       "Choose from tier list: surface / pattern / structural / causal / theoretical"),
        ("base_bank_depth",      75,         "0 – 100 (int)",  "Knowledge-base depth score (informational — not used in rarity formula)"),
        ("inferential_richness", 0.82,       "0.00 – 1.00",    "Depth of reasoning chain (0 = shallow, 1 = maximal depth)"),
        ("trend_position",       0.35,       "0.00 – 1.00",    "0 = emerging edge (rare), 1 = well-established (common)"),
        ("issuance_trigger",     "causal_chain_identified", "free text", "Cognitive event that triggered this issuance"),
    ]
    for i, (field, default, allowed, note) in enumerate(inputs_def):
        row = 6 + i
        lbl(ws, row, 1, field, bold=False)
        inp(ws, row, 2, default)
        w(ws, row, 3, allowed if isinstance(allowed, str) else " / ".join(allowed),
          font=_f(BLACK_TXT, sz=9, italic=True), align=_left())
        w(ws, row, 4, note, font=_f(BLACK_TXT, sz=9), align=_wrap(), border=_border())
        ws.row_dimensions[row].height = 16

    # tier data validation
    dv = DataValidation(type="list", formula1='"surface,pattern,structural,causal,theoretical"', allow_blank=False)
    dv.sqref = "B6"
    ws.add_data_validation(dv)

    # ── Computed Outputs ──────────────────────────────────────────────────────
    sect(ws, 13, 1, "Computed Outputs", span=4)
    hdr_row(ws, 14, ["Output", "Value", "Formula", "Notes"])

    # Tier rarity weight lookup: VLOOKUP(B6, Assumptions!$A$10:$B$14, 2, FALSE)
    # Emergence component: Assumptions!$B$3 * (1 - B9)
    # Inferential component: Assumptions!$B$4 * B8
    # Tier component: Assumptions!$B$5 * VLOOKUP(...)
    # Rarity score: sum of three components, clamped 0-1

    outputs = [
        ("tier_rarity_weight",
         '=IFERROR(VLOOKUP(B6,Assumptions!$A$10:$B$14,2,FALSE),"N/A")',
         '=VLOOKUP(B6, Assumptions!A10:B14, 2, FALSE)',
         "Tier-specific weight from Assumptions table"),
        ("tier_amount",
         '=IFERROR(VLOOKUP(B6,Assumptions!$A$10:$C$14,3,FALSE),"N/A")',
         '=VLOOKUP(B6, Assumptions!A10:C14, 3, FALSE)',
         "Backward-compatible integer amount for this tier"),
        ("emergence_component",
         '=IFERROR(Assumptions!$B$3*(1-B9),"N/A")',
         '=w_emergence × (1 − trend_position)',
         "40% weight on how far from well-established edge"),
        ("inferential_component",
         '=IFERROR(Assumptions!$B$4*B8,"N/A")',
         '=w_inferential × inferential_richness',
         "40% weight on depth of reasoning chain"),
        ("tier_component",
         '=IFERROR(Assumptions!$B$5*VLOOKUP(B6,Assumptions!$A$10:$B$14,2,FALSE),"N/A")',
         '=w_tier × tier_rarity_weight',
         "20% weight on epistemic tier"),
        ("rarity_score  (0.00 – 1.00)",
         '=IFERROR(MAX(0,MIN(1,B17+B18+B19)),"N/A")',
         '=MAX(0, MIN(1, emergence + inferential + tier))',
         "Final rarity score stamped at issuance — NEVER recompute after issue"),
    ]
    for i, (out_name, formula, formula_text, note) in enumerate(outputs):
        row = 15 + i
        lbl(ws, row, 1, out_name, bold=(i == 5))
        form(ws, row, 2, formula, fmt="0.0000" if i not in (1,) else "0",
             green=(i == 5))
        w(ws, row, 3, formula_text, font=_f(BLACK_TXT, sz=9, italic=True), align=_wrap())
        w(ws, row, 4, note, font=_f(BLACK_TXT, sz=9), align=_wrap(), border=_border())
        ws.row_dimensions[row].height = 18
        if i == 5:
            ws.cell(row=row, column=2).fill = _fill(PASS_BG)

    # ── Rarity interpretation ─────────────────────────────────────────────────
    sect(ws, 23, 1, "Rarity Score Interpretation Guide", span=4)
    hdr_row(ws, 24, ["Score Range", "Classification", "Meaning", "Typical Tier"])
    interp = [
        ("0.00 – 0.25", "Common",     "Well-established knowledge, shallow reasoning",  "surface / pattern"),
        ("0.25 – 0.50", "Uncommon",   "Moderate depth, some emerging edge",              "pattern / structural"),
        ("0.50 – 0.75", "Rare",       "High inferential depth, emerging domain",         "structural / causal"),
        ("0.75 – 1.00", "Very Rare",  "Theoretical synthesis at knowledge frontier",     "causal / theoretical"),
    ]
    for i, row_data in enumerate(interp):
        row = 25 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        for j, val in enumerate(row_data):
            c = w(ws, row, 1+j, val, font=_f(BLACK_TXT), align=_left(), border=_border())
            if row_fill: c.fill = row_fill

    col_w(ws, {"A": 30, "B": 18, "C": 38, "D": 50})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Tier_Ref
# ═══════════════════════════════════════════════════════════════════════════════
def build_tier_ref(wb):
    ws = wb.create_sheet("Tier_Ref")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Insight Tier Reference  (domain.py: InsightTier, lines 48–78)", span=6)
    hdr_row(ws, 2, ["Tier Key", "Order", "Amount (int)", "Rarity Weight", "Description", "Cognitive Demand"])

    tier_extended = [
        ("surface",     1, 1, 0.00, "Fact recall, basic recognition",                    "Low — retrieval and identification"),
        ("pattern",     2, 2, 0.25, "Cross-domain pattern recognition",                   "Moderate — synthesis across contexts"),
        ("structural",  3, 3, 0.50, "Structural understanding, system maps",              "High — system-level comprehension"),
        ("causal",      4, 4, 0.75, "Causal chain modelling, root cause analysis",        "Very High — causal inference & modelling"),
        ("theoretical", 5, 5, 1.00, "Novel synthesis, theory building at frontier",       "Maximal — new theory generation"),
    ]

    fills = [_fill("FFE2EFDA"), _fill("FFEBF7E2"), _fill("FFFFE699"),
             _fill("FFFFC7CE"), _fill("FFE2CFFF")]

    for i, (tier, order, amt, rw, desc, demand) in enumerate(tier_extended):
        row = 3 + i
        ws.row_dimensions[row].height = 22
        # tier key  — link style to Assumptions
        c = w(ws, row, 1, f'=Assumptions!A{10+i}',
              font=_f(GREEN_TXT, bold=True), fill=fills[i],
              align=_ctr(), border=_border())
        w(ws, row, 2, f'=Assumptions!E{10+i}',
          font=_f(BLACK_TXT), fill=fills[i], align=_ctr(), border=_border(), fmt="0")
        w(ws, row, 3, f'=Assumptions!C{10+i}',
          font=_f(BLACK_TXT), fill=fills[i], align=_ctr(), border=_border(), fmt="0")
        w(ws, row, 4, f'=Assumptions!B{10+i}',
          font=_f(BLACK_TXT), fill=fills[i], align=_ctr(), border=_border(), fmt="0.00")
        w(ws, row, 5, desc,
          font=_f(BLACK_TXT), fill=fills[i], align=_wrap(), border=_border())
        w(ws, row, 6, demand,
          font=_f(BLACK_TXT, sz=9), fill=fills[i], align=_wrap(), border=_border())

    # Rarity formula note
    ws.row_dimensions[10].height = 8
    sect(ws, 11, 1, "Rarity Formula  (domain.py:compute_rarity_score)", span=6)
    formula_text = (
        "rarity_score = "
        "w_emergence × (1 − trend_position)  +  "
        "w_inferential × inferential_richness  +  "
        "w_tier × tier_rarity_weight\n"
        "   where  w_emergence=0.4,  w_inferential=0.4,  w_tier=0.2  (see Assumptions!B3:B5)\n"
        "Result clamped to [0.0, 1.0] and rounded to 4 decimal places."
    )
    ws.merge_cells("A12:F14")
    c = ws.cell(row=12, column=1, value=formula_text)
    c.font = _f(BLACK_TXT, sz=SZ)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = _border()
    ws.row_dimensions[12].height = 60

    col_w(ws, {"A": 15, "B": 8, "C": 15, "D": 15, "E": 42, "F": 40})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Lifecycle
# ═══════════════════════════════════════════════════════════════════════════════
def build_lifecycle(wb):
    ws = wb.create_sheet("Lifecycle")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Reward State Machine  (domain.py: RewardState, lines 14–22)", span=7)

    # ── A. States ─────────────────────────────────────────────────────────────
    sect(ws, 3, 1, "A. State Definitions", span=3)
    hdr_row(ws, 4, ["State", "Outcome State", "Meaning"])
    state_rows = [
        ("drafted",              "unknown",              "Reward created; awaiting contract satisfaction"),
        ("earned",               "unknown",              "Contract satisfied; awaiting payment instruction"),
        ("payment_pending",      "unknown",              "Ready for payment; Stripe charge not yet confirmed"),
        ("payment_confirmed",    "delivered_pending_ack","Stripe confirmed; awaiting student acknowledgement"),
        ("student_acknowledged", "acknowledged",         "Student acknowledged receipt — terminal success state"),
        ("review_requested",     "review_open",          "Review flag raised from any state — lateral transition"),
    ]
    state_fills = {
        "drafted":              _fill("FFF2F2F2"),
        "earned":               _fill("FFE2EFDA"),
        "payment_pending":      _fill("FFFFE699"),
        "payment_confirmed":    _fill("FFBDD7EE"),
        "student_acknowledged": _fill("FFC6EFCE"),
        "review_requested":     _fill("FFFFC7CE"),
    }
    for i, (state, outcome, meaning) in enumerate(state_rows):
        row = 5 + i
        ws.row_dimensions[row].height = 20
        sf = state_fills[state]
        w(ws, row, 1, state,   font=_f(BLACK_TXT, bold=True), fill=sf, align=_ctr(), border=_border())
        w(ws, row, 2, outcome, font=_f(BLACK_TXT),             fill=sf, align=_left(), border=_border())
        w(ws, row, 3, meaning, font=_f(BLACK_TXT, sz=9),       fill=sf, align=_wrap(), border=_border())

    # ── B. Valid Transitions (from → to matrix) ───────────────────────────────
    sect(ws, 13, 1, "B. Valid Transition Matrix", span=8)
    hdr_row(ws, 14, ["From State", "→ To State", "Trigger", "Policy Flag", "Notes"], start_col=1)
    transitions = [
        ("drafted",              "earned",               "glass_session_event",  "contract_satisfied=true",  "domain.py:380-384"),
        ("earned",               "payment_pending",       "glass_session_event",  "ready_for_payment=true",   "domain.py:386-389"),
        ("earned",               "payment_confirmed",     "stripe_payment_success","student_id matches",      "domain.py:407-440"),
        ("payment_pending",      "payment_confirmed",     "stripe_payment_success","student_id matches",      "domain.py:407-440"),
        ("payment_confirmed",    "student_acknowledged",  "student_confirmation", "student_ack=true",         "domain.py:364-375"),
        ("any",                  "review_requested",      "any evidence",         "request_review=true",      "domain.py:357-362"),
        ("payment_confirmed",    "payment_confirmed",     "stripe_payment_success","replay — already confirmed","domain.py:419-424"),
        ("student_acknowledged", "student_acknowledged",  "stripe_payment_success","replay — already confirmed","domain.py:419-424"),
    ]
    for i, (frm, to, trigger, flag, src) in enumerate(transitions):
        row = 15 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        vals = [frm, to, trigger, flag, src]
        for j, val in enumerate(vals):
            c = w(ws, row, 1+j, val, font=_f(BLACK_TXT, sz=9), align=_left(), border=_border())
            if row_fill: c.fill = row_fill

    # ── C. Evidence Types ─────────────────────────────────────────────────────
    sect(ws, 25, 1, "C. Evidence Types  (domain.py:EvidenceType)", span=3)
    hdr_row(ws, 26, ["Evidence Type", "Triggers Transition?", "Notes"])
    ev_rows = [
        ("glass_session_event",  "Only if policy flags set",    "Most common; drives drafted→earned→payment_pending"),
        ("failure_snapshot",     "Notes failure; no state change","Recorded in notes; domain.py:391-393"),
        ("return_attempt",       "No (returns None)",           "Non-transition evidence type"),
        ("success_after_support","No (returns None)",           "Non-transition evidence type"),
        ("agent_interpretation", "No (returns None)",           "Non-transition evidence type"),
        ("student_confirmation", "Yes → student_acknowledged",  "Requires current=payment_confirmed"),
    ]
    for i, (ev, trig, note) in enumerate(ev_rows):
        row = 27 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        for j, val in enumerate([ev, trig, note]):
            c = w(ws, row, 1+j, val, font=_f(BLACK_TXT, sz=9), align=_wrap(), border=_border())
            ws.row_dimensions[row].height = 18
            if row_fill: c.fill = row_fill

    col_w(ws, {"A": 26, "B": 26, "C": 30, "D": 30, "E": 30})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Token_Log
# ═══════════════════════════════════════════════════════════════════════════════
def build_token_log(wb):
    ws = wb.create_sheet("Token_Log")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Token Issuance Portfolio Log  (15 sample tokens)", span=9)
    w(ws, 2, 1,
      "Rarity Score and Amount columns are computed via formulas referencing Assumptions sheet. "
      "Blue-text columns are editable. Black-text columns are formula outputs.",
      font=_f(BLACK_TXT, italic=True, sz=9), align=_wrap())
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:I2")

    headers = [
        "Token ID", "Issued At", "Insight Tier",
        "Base Bank Depth\n(0–100)", "Inferential\nRichness (0-1)",
        "Trend Position\n(0=emerging, 1=common)",
        "Rarity Score\n(formula)", "Amount\n(formula)", "Issuance Trigger",
    ]
    hdr_row(ws, 3, headers)
    ws.row_dimensions[3].height = 32

    # 15 sample tokens
    tokens = [
        ("TKN-001", "2026-04-01T09:15:00Z", "theoretical",  88, 0.91, 0.12, "novel_synthesis_identified"),
        ("TKN-002", "2026-04-03T14:22:00Z", "causal",       76, 0.83, 0.28, "causal_chain_identified"),
        ("TKN-003", "2026-04-05T11:05:00Z", "structural",   65, 0.72, 0.45, "system_map_constructed"),
        ("TKN-004", "2026-04-07T16:40:00Z", "pattern",      55, 0.58, 0.62, "cross_domain_link_found"),
        ("TKN-005", "2026-04-10T08:30:00Z", "surface",      42, 0.22, 0.88, "fact_recalled_correctly"),
        ("TKN-006", "2026-04-12T13:55:00Z", "theoretical",  92, 0.95, 0.08, "theory_extension_proposed"),
        ("TKN-007", "2026-04-14T10:10:00Z", "causal",       81, 0.78, 0.33, "root_cause_identified"),
        ("TKN-008", "2026-04-17T15:25:00Z", "structural",   70, 0.68, 0.51, "dependency_graph_completed"),
        ("TKN-009", "2026-04-20T09:45:00Z", "causal",       73, 0.85, 0.22, "mechanism_traced_to_source"),
        ("TKN-010", "2026-04-22T12:00:00Z", "pattern",      60, 0.55, 0.70, "recurring_pattern_named"),
        ("TKN-011", "2026-04-25T14:35:00Z", "theoretical",  95, 0.93, 0.05, "conceptual_bridge_built"),
        ("TKN-012", "2026-04-28T11:20:00Z", "structural",   68, 0.65, 0.48, "abstraction_layer_identified"),
        ("TKN-013", "2026-05-01T09:00:00Z", "surface",      38, 0.18, 0.92, "definition_confirmed"),
        ("TKN-014", "2026-05-03T16:10:00Z", "causal",       79, 0.80, 0.30, "feedback_loop_identified"),
        ("TKN-015", "2026-05-05T10:50:00Z", "pattern",      58, 0.62, 0.58, "analogical_transfer_made"),
    ]

    for i, (tid, issued, tier, depth, ir, tp, trigger) in enumerate(tokens):
        row = 4 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None

        # Inputs (blue)
        inp(ws, row, 1, tid)
        inp(ws, row, 2, issued)
        inp(ws, row, 3, tier)
        inp(ws, row, 4, depth,  fmt="0")
        inp(ws, row, 5, ir,     fmt="0.00")
        inp(ws, row, 6, tp,     fmt="0.00")
        inp(ws, row, 9, trigger)

        # Rarity score formula
        # =MAX(0,MIN(1, Assumptions!$B$3*(1-F{row}) + Assumptions!$B$4*E{row}
        #               + Assumptions!$B$5*VLOOKUP(C{row},Assumptions!$A$10:$B$14,2,FALSE)))
        rarity_f = (
            f"=IFERROR(MAX(0,MIN(1,"
            f"Assumptions!$B$3*(1-F{row})"
            f"+Assumptions!$B$4*E{row}"
            f"+Assumptions!$B$5*VLOOKUP(C{row},Assumptions!$A$10:$B$14,2,FALSE)"
            f")),\"ERR\")"
        )
        amount_f = f'=IFERROR(VLOOKUP(C{row},Assumptions!$A$10:$C$14,3,FALSE),"ERR")'

        c_r = form(ws, row, 7, rarity_f, fmt="0.0000", green=True)
        c_a = form(ws, row, 8, amount_f, fmt="0",       green=True)

        if row_fill:
            c_r.fill = row_fill
            c_a.fill = row_fill

        ws.row_dimensions[row].height = 16

    # data validation for tier column
    dv = DataValidation(type="list", formula1='"surface,pattern,structural,causal,theoretical"')
    dv.sqref = f"C4:C18"
    ws.add_data_validation(dv)

    # ── Portfolio Statistics ───────────────────────────────────────────────────
    stat_row = 21
    sect(ws, stat_row, 1, "Portfolio Statistics", span=4)
    hdr_row(ws, stat_row + 1, ["Metric", "Value", "Formula Note", ""])

    stats = [
        ("Token Count",          f"=COUNTA(A4:A18)",                                            "Count all non-empty token IDs"),
        ("Total Amount",         f"=IFERROR(SUM(H4:H18),\"ERR\")",                               "Sum of all token amounts"),
        ("Average Rarity Score", f"=IFERROR(AVERAGE(G4:G18),\"ERR\")",                           "Mean rarity across portfolio"),
        ("Max Rarity Score",     f"=IFERROR(MAX(G4:G18),\"ERR\")",                               "Highest rarity in portfolio"),
        ("Min Rarity Score",     f"=IFERROR(MIN(G4:G18),\"ERR\")",                               "Lowest rarity in portfolio"),
        ("Theoretical Tokens",   f'=COUNTIF(C4:C18,"theoretical")',                               "Count of theoretical-tier tokens"),
        ("Causal Tokens",        f'=COUNTIF(C4:C18,"causal")',                                    "Count of causal-tier tokens"),
        ("Structural Tokens",    f'=COUNTIF(C4:C18,"structural")',                                "Count of structural-tier tokens"),
        ("Pattern Tokens",       f'=COUNTIF(C4:C18,"pattern")',                                   "Count of pattern-tier tokens"),
        ("Surface Tokens",       f'=COUNTIF(C4:C18,"surface")',                                   "Count of surface-tier tokens"),
        ("Very Rare (≥0.75)",    f'=COUNTIF(G4:G18,">="&0.75)',                                   "Rarity ≥ 0.75"),
        ("Rare (0.50–0.74)",     f'=COUNTIFS(G4:G18,">="&0.5,G4:G18,"<"&0.75)',                  "0.50 ≤ rarity < 0.75"),
        ("Uncommon (0.25–0.49)", f'=COUNTIFS(G4:G18,">="&0.25,G4:G18,"<"&0.5)',                  "0.25 ≤ rarity < 0.50"),
        ("Common (<0.25)",       f'=COUNTIF(G4:G18,"<"&0.25)',                                    "Rarity < 0.25"),
    ]
    for i, (metric, formula, note) in enumerate(stats):
        row = stat_row + 2 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        lbl(ws, row, 1, metric, bold=False)
        c = form(ws, row, 2, formula, fmt=("0.0000" if "Rarity" in metric and "Count" not in metric else "0"))
        w(ws, row, 3, note, font=_f(BLACK_TXT, sz=9, italic=True), align=_left())
        if row_fill: c.fill = row_fill

    col_w(ws, {"A": 14, "B": 22, "C": 10, "D": 12,
               "E": 14, "F": 16, "G": 14, "H": 12, "I": 28})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Exchange_Eval
# ═══════════════════════════════════════════════════════════════════════════════
def build_exchange_eval(wb):
    ws = wb.create_sheet("Exchange_Eval")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Exchange Constraint Evaluator  (domain.py:evaluate_exchange_request)", span=6)
    w(ws, 2, 1,
      "Enter the requested scope (up to 10 key-value pairs) and the approval flag. "
      "The three constraint layers evaluate in order: Safety → Ethics → Economic Stability.",
      font=_f(BLACK_TXT, italic=True, sz=9), align=_wrap())
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:F2")

    # ── Request Inputs ────────────────────────────────────────────────────────
    sect(ws, 4, 1, "Request Inputs", span=6)
    lbl(ws, 5, 1, "Request ID")
    inp(ws, 5, 2, "REQ-001")
    lbl(ws, 6, 1, "Student ID")
    inp(ws, 6, 2, "STU-042")
    lbl(ws, 7, 1, "Reward ID")
    inp(ws, 7, 2, "RWD-007")
    lbl(ws, 8, 1, "Explicit Irreversible Approval (TRUE/FALSE)")
    inp(ws, 8, 2, "FALSE")

    sect(ws, 10, 1, "Requested Scope  (up to 10 key-value pairs)", span=6)
    hdr_row(ws, 11, ["Scope Key", "Scope Value", "Type / Notes", "", "", ""])
    scope_inputs = [
        ("token_amount",        3,         "int — capped by max_token_amount"),
        ("knowledge_domain",    "physics", "str — domain of knowledge exchange"),
        ("session_access",      "TRUE",    "bool — access to session data"),
        ("comment",             "earned_via_causal_chain", "str — human notes"),
        ("reward_id",           "RWD-007", "str — associated reward"),
        ("",                    "",        ""),
        ("",                    "",        ""),
        ("",                    "",        ""),
        ("",                    "",        ""),
        ("",                    "",        ""),
    ]
    for i, (key, val, note) in enumerate(scope_inputs):
        row = 12 + i
        inp(ws, row, 1, key)
        inp(ws, row, 2, val)
        w(ws, row, 3, note, font=_f(BLACK_TXT, sz=9, italic=True), align=_left(), border=_border())

    # ── Layer 1: Safety / Security ────────────────────────────────────────────
    sect(ws, 24, 1, "Layer 1: Safety / Security  (Hard Gate)", span=6)
    hdr_row(ws, 25, ["Scope Key", "In Blocked List?", "Layer 1 Result", "", "", ""])
    for i in range(10):
        row = 26 + i
        key_ref = f"A{12+i}"
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        c1 = form(ws, row, 1, f"={key_ref}", green=True)
        # Check if key is in blocked list (Assumptions!$A$25:$A$37)
        blocked_check = (
            f'=IF({key_ref}="","—",'
            f'IF(COUNTIF(Assumptions!$A$25:$A$37,{key_ref})>0,"BLOCKED","PASS"))'
        )
        c2 = form(ws, row, 2, blocked_check)
        # Colour
        if row_fill:
            c1.fill = row_fill
            c2.fill = row_fill

    # Overall Layer 1 decision
    lbl(ws, 37, 1, "Layer 1 Overall Decision")
    l1_decision = (
        '=IF(COUNTIF(B26:B35,"BLOCKED")>0,'
        '"✗ REJECTED — blocked scope key present","✓ PASSED")'
    )
    c = form(ws, 37, 2, l1_decision)
    c.font = _f(BLACK_TXT, bold=True)

    lbl(ws, 38, 1, "Blocked Key Count")
    form(ws, 38, 2, '=COUNTIF(B26:B35,"BLOCKED")', fmt="0")

    # ── Layer 2: Ethics / Irreversibility ─────────────────────────────────────
    sect(ws, 40, 1, "Layer 2: Ethics / Irreversibility  (Scope Narrowing)", span=6)
    hdr_row(ws, 41, ["Scope Key", "Irreversible?", "Approval Flag", "Layer 2 Result", "", ""])
    for i in range(10):
        row = 42 + i
        key_ref = f"A{12+i}"
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        c1 = form(ws, row, 1, f"={key_ref}", green=True)
        # Check if key is in irreversible list (Assumptions!$A$36:$A$47)
        irrev_check = (
            f'=IF({key_ref}="","—",'
            f'IF(COUNTIF(Assumptions!$A$36:$A$47,{key_ref})>0,"YES","NO"))'
        )
        c2 = form(ws, row, 2, irrev_check)
        c3 = form(ws, row, 3, "=$B$8")  # approval flag
        # Layer 2 result: if irreversible AND require approval AND no approval → STRIPPED
        layer2_result = (
            f'=IF({key_ref}="","—",'
            f'IF(AND(COUNTIF(Assumptions!$A$36:$A$47,{key_ref})>0,'
            f'Assumptions!$B$21="TRUE",$B$8="FALSE"),'
            f'"STRIPPED","APPROVED"))'
        )
        c4 = form(ws, row, 4, layer2_result)
        if row_fill:
            for c in (c1, c2, c3, c4): c.fill = row_fill

    lbl(ws, 53, 1, "Layer 2 Keys Stripped")
    form(ws, 53, 2, '=COUNTIF(D42:D51,"STRIPPED")', fmt="0")
    lbl(ws, 54, 1, "Layer 2 Overall Status")
    form(ws, 54, 2, '=IF(D37="✗ REJECTED — blocked scope key present","N/A (Layer 1 rejected)","✓ PASSED")')

    # ── Layer 3: Economic Stability ───────────────────────────────────────────
    sect(ws, 56, 1, "Layer 3: Economic Stability  (Scope Narrowing)", span=6)
    hdr_row(ws, 57, ["Scope Key", "Scope Value", "Layer 3 Result", "Capped Value", "", ""])

    # Count approved scope items (passed L1 and not stripped in L2)
    approved_count_f = (
        '=SUMPRODUCT((A12:A21<>"")*(B26:B35<>"BLOCKED")*(D42:D51<>"STRIPPED"))'
    )
    lbl(ws, 58, 1, "Approved Scope Item Count")
    form(ws, 58, 2, approved_count_f, fmt="0")
    lbl(ws, 59, 1, "Max Scope Items")
    form(ws, 59, 2, "=Assumptions!$B$20", fmt="0", green=True)
    lbl(ws, 60, 1, "Max Token Amount")
    form(ws, 60, 2, "=Assumptions!$B$19", fmt="0", green=True)

    for i in range(10):
        row = 62 + i
        key_ref    = f"A{12+i}"
        val_ref    = f"B{12+i}"
        l1_ref     = f"B{26+i}"
        l2_ref     = f"D{42+i}"
        row_fill   = _fill(ALT_BG) if i % 2 == 0 else None
        c1 = form(ws, row, 1, f"={key_ref}", green=True)
        c2 = form(ws, row, 2, f"={val_ref}", green=True)

        # Layer 3 result: approved if L1 passed, L2 approved, and within scope count
        # (rank within approved items — simplified: approved if order ≤ max_scope_items)
        layer3_result = (
            f'=IF({key_ref}="","—",'
            f'IF({l1_ref}="BLOCKED","BLOCKED (L1)",'
            f'IF({l2_ref}="STRIPPED","STRIPPED (L2)",'
            f'IF(SUMPRODUCT(({l1_ref}<>"BLOCKED")*({l2_ref}<>"STRIPPED")*(A12:A{11+i}<>"")*(B26:B{25+i}<>"BLOCKED")*(D42:D{41+i}<>"STRIPPED"))+1>$B$59,'
            f'"OVER_LIMIT","APPROVED"))))'
        )
        c3 = form(ws, row, 3, layer3_result)

        # token_amount cap
        capped = (
            f'=IF(AND({key_ref}="token_amount",'
            f'ISNUMBER({val_ref}),{val_ref}>$B$60),'
            f'$B$60,{val_ref})'
        )
        c4 = form(ws, row, 4, capped)
        if row_fill:
            for c in (c1, c2, c3, c4): c.fill = row_fill

    # ── Final Result ──────────────────────────────────────────────────────────
    sect(ws, 74, 1, "Final Exchange Result", span=6)
    hdr_row(ws, 75, ["Field", "Value", "", "", "", ""])
    results = [
        ("Overall Approved",
         '=IF(B37="✓ PASSED","TRUE","FALSE")'),
        ("Items in Final Approved Scope",
         '=COUNTIF(C62:C71,"APPROVED")'),
        ("Items Blocked (Safety)",
         '=COUNTIF(C62:C71,"BLOCKED (L1)")'),
        ("Items Stripped (Ethics)",
         '=COUNTIF(C62:C71,"STRIPPED (L2)")'),
        ("Items Dropped (Over Limit)",
         '=COUNTIF(C62:C71,"OVER_LIMIT")'),
        ("token_amount After Cap",
         '=IFERROR(INDEX(D62:D71,MATCH("token_amount",A12:A21,0)),"N/A")'),
    ]
    for i, (metric, formula) in enumerate(results):
        row = 76 + i
        lbl(ws, row, 1, metric, bold=False)
        c = form(ws, row, 2, formula, fmt=("0" if i > 0 else None))
        if i == 0:
            c.font = _f(BLACK_TXT, bold=True)

    col_w(ws, {"A": 40, "B": 22, "C": 18, "D": 16, "E": 12, "F": 12})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_summary(wb):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    sect(ws, 1, 1, "Portfolio Summary Dashboard  (draws from Token_Log)", span=6)
    w(ws, 2, 1,
      "All values computed from Token_Log via formulas. No manual inputs on this sheet.",
      font=_f(BLACK_TXT, italic=True, sz=9), align=_wrap())
    ws.merge_cells("A2:F2")

    # ── A. Portfolio Overview ─────────────────────────────────────────────────
    sect(ws, 4, 1, "A. Portfolio Overview", span=6)
    hdr_row(ws, 5, ["Metric", "Value", "", "", "", ""])
    overview = [
        ("Total Tokens Issued",   "=COUNTA(Token_Log!A4:A18)",           "0"),
        ("Total Token Amount",    "=IFERROR(SUM(Token_Log!H4:H18),0)",    "0"),
        ("Average Rarity Score",  "=IFERROR(AVERAGE(Token_Log!G4:G18),0)","0.0000"),
        ("Median Rarity Score",   "=IFERROR(MEDIAN(Token_Log!G4:G18),0)", "0.0000"),
        ("Max Rarity Score",      "=IFERROR(MAX(Token_Log!G4:G18),0)",    "0.0000"),
        ("Min Rarity Score",      "=IFERROR(MIN(Token_Log!G4:G18),0)",    "0.0000"),
        ("Std Dev Rarity",        "=IFERROR(STDEV(Token_Log!G4:G18),0)",  "0.0000"),
    ]
    for i, (metric, formula, fmt) in enumerate(overview):
        row = 6 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        lbl(ws, row, 1, metric, bold=False)
        c = form(ws, row, 2, formula, fmt=fmt, green=True)
        if row_fill: c.fill = row_fill

    # ── B. Tokens by Tier ─────────────────────────────────────────────────────
    sect(ws, 15, 1, "B. Tokens by Tier", span=6)
    hdr_row(ws, 16, ["Insight Tier", "Count", "% of Portfolio", "Avg Rarity", "Avg Amount", "Total Amount"])
    for i, (tier, _, amt, _) in enumerate(TIERS):
        row = 17 + i
        row_fill = _fill(ALT_BG) if i % 2 == 0 else None
        lbl(ws, row, 1, tier, bold=False)
        count_f   = f'=COUNTIF(Token_Log!C4:C18,"{tier}")'
        pct_f     = f'=IFERROR(B{row}/SUM(B17:B21),0)'
        avg_rar_f = f'=IFERROR(AVERAGEIF(Token_Log!C4:C18,"{tier}",Token_Log!G4:G18),"-")'
        avg_amt_f = f'=IFERROR(AVERAGEIF(Token_Log!C4:C18,"{tier}",Token_Log!H4:H18),"-")'
        tot_amt_f = f'=IFERROR(SUMIF(Token_Log!C4:C18,"{tier}",Token_Log!H4:H18),0)'
        vals = [(count_f,"0"),(pct_f,"0.0%"),(avg_rar_f,"0.0000"),(avg_amt_f,"0.0"),(tot_amt_f,"0")]
        for j, (f, fmt) in enumerate(vals):
            c = form(ws, row, 2+j, f, fmt=fmt, green=True)
            if row_fill: c.fill = row_fill

    # Totals row
    row = 22
    lbl(ws, row, 1, "TOTAL", bold=True)
    for j, (f, fmt) in enumerate([
        ("=SUM(B17:B21)","0"),
        ("=SUM(C17:C21)","0.0%"),
        ('=IFERROR(AVERAGE(Token_Log!G4:G18),"")','0.0000'),
        ('=IFERROR(AVERAGE(Token_Log!H4:H18),"")','0.0'),
        ("=SUM(F17:F21)","0"),
    ]):
        c = form(ws, row, 2+j, f, fmt=fmt, green=True)
        c.font = _f(BLACK_TXT, bold=True)

    # ── C. Rarity Band Distribution ───────────────────────────────────────────
    sect(ws, 24, 1, "C. Rarity Band Distribution", span=6)
    hdr_row(ws, 25, ["Band", "Score Range", "Count", "% of Portfolio", "Avg Amount", ""])
    bands = [
        ("Very Rare",  "≥ 0.75", '=COUNTIF(Token_Log!G4:G18,">="&0.75)',                               '=IFERROR(AVERAGEIF(Token_Log!G4:G18,">="&0.75,Token_Log!H4:H18),"-")'),
        ("Rare",       "0.50–0.74", '=COUNTIFS(Token_Log!G4:G18,">="&0.5,Token_Log!G4:G18,"<"&0.75)', '=IFERROR(AVERAGEIFS(Token_Log!H4:H18,Token_Log!G4:G18,">="&0.5,Token_Log!G4:G18,"<"&0.75),"-")'),
        ("Uncommon",   "0.25–0.49", '=COUNTIFS(Token_Log!G4:G18,">="&0.25,Token_Log!G4:G18,"<"&0.5)', '=IFERROR(AVERAGEIFS(Token_Log!H4:H18,Token_Log!G4:G18,">="&0.25,Token_Log!G4:G18,"<"&0.5),"-")'),
        ("Common",     "< 0.25",   '=COUNTIF(Token_Log!G4:G18,"<"&0.25)',                              '=IFERROR(AVERAGEIF(Token_Log!G4:G18,"<"&0.25,Token_Log!H4:H18),"-")'),
    ]
    band_fills = [_fill("FFFFE699"), _fill("FFBDD7EE"), _fill("FFE2EFDA"), _fill("FFF2F2F2")]
    for i, (band, rng, count_f, avg_f) in enumerate(bands):
        row = 26 + i
        pct_f = f"=IFERROR(C{row}/SUM(C26:C29),0)"
        lbl(ws, row, 1, band, bold=False)
        for j, (f, fmt) in enumerate([(rng,None),(count_f,"0"),(pct_f,"0.0%"),(avg_f,"0.0")]):
            if j == 0:
                w(ws, row, 2, f, font=_f(BLACK_TXT, sz=9, italic=True),
                  align=_ctr(), border=_border(), fill=band_fills[i])
            else:
                c = form(ws, row, 2+j, f, fmt=fmt, green=(j>0))
                c.fill = band_fills[i]

    col_w(ws, {"A": 26, "B": 16, "C": 10, "D": 14, "E": 14, "F": 14})


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    build_cover(wb)
    build_assumptions(wb)
    build_rarity_calc(wb)
    build_tier_ref(wb)
    build_lifecycle(wb)
    build_token_log(wb)
    build_exchange_eval(wb)
    build_summary(wb)

    out = "/tmp/opencode/xchange_reward_token_calculator.xlsx"
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
